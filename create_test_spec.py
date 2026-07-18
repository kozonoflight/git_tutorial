#!/usr/bin/env python3
"""Primo dataLayer テスト仕様書 Excel生成 v2"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="2E86AB")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
NORMAL = Font(size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

SITE = "primo-demo2-front.ec-rider2-demo.net"
GTM_ID = "GTM-TGF9D8WX"

# (ID, カテゴリ, テスト項目, 対象, 優先度, 種別)
TEST_CASES = [
    # GTM
    ("TC-001", "GTM基本設定", "GTMスニペット（head）の設置", "GTM head", "高", "正常"),
    ("TC-002", "GTM基本設定", "GTMスニペット（body noscript）の設置", "GTM body", "高", "正常"),
    ("TC-003", "GTM基本設定", "User-ID初期化がGTMスニペットより前に実行される", "実装順序", "高", "正常"),
    ("TC-004", "GTM基本設定", "dataLayer初期化処理", "dataLayer init", "高", "正常"),
    ("TC-005", "GTM基本設定", "GTMスニペットが1回のみ設置されている", "GTM head", "高", "正常"),
    # COMMON
    ("TC-010", "共通ページ計測", "ログイン時：user_id/area_id/staff_id/member_r送信", "共通push", "高", "正常"),
    ("TC-011", "共通ページ計測", "未ログイン時：user_idキーを出力しない", "共通push", "高", "正常"),
    ("TC-012", "共通ページ計測", "member_rの数値マッピング（0〜3）", "member_r", "中", "正常"),
    ("TC-013", "共通ページ計測", "ページリロード時のdataLayer再出力", "共通push", "中", "正常"),
    ("TC-014", "共通ページ計測", "通常ページ表示時にloginイベントを送信しない", "login", "中", "異常"),
    # AUTH
    ("TC-020", "認証", "ログイン成功時にloginイベント送信", "login", "高", "正常"),
    ("TC-021", "認証", "ログイン失敗時はloginイベントを送信しない", "login", "高", "異常"),
    ("TC-022", "認証", "会員登録成功時にsign_upイベント送信", "sign_up", "高", "正常"),
    ("TC-023", "認証", "会員登録失敗時はsign_upイベントを送信しない", "sign_up", "高", "異常"),
    ("TC-024", "認証", "会員登録成功時：mailmagazine_signup（subscription_status=all）", "mailmagazine_signup", "高", "正常"),
    ("TC-025", "認証", "会員登録成功時：mailmagazine_signup（subscription_status=text_only）", "mailmagazine_signup", "中", "正常"),
    ("TC-026", "認証", "会員登録成功時：mailmagazine_signup（subscription_status=none）", "mailmagazine_signup", "中", "正常"),
    ("TC-027", "認証", "会員登録失敗時はmailmagazine_signupを送信しない", "mailmagazine_signup", "高", "異常"),
    # LEAD
    ("TC-030", "リード", "見積書DL時にquote_downloadイベント送信", "quote_download", "高", "正常"),
    ("TC-031", "リード", "見積依頼成功時：generate_lead（lead_source=quote）+items", "generate_lead", "高", "正常"),
    ("TC-032", "リード", "問い合わせ成功時：generate_lead（lead_source=inquiry）", "generate_lead", "高", "正常"),
    ("TC-033", "リード", "再見積（re_quote）時はgenerate_leadを送信しない", "generate_lead", "高", "異常"),
    ("TC-034", "リード", "見積依頼バリデーションエラー時はgenerate_leadを送信しない", "generate_lead", "高", "異常"),
    ("TC-035", "リード", "問い合わせ送信失敗時はgenerate_leadを送信しない", "generate_lead", "高", "異常"),
    # ECOM
    ("TC-040", "商品・カート", "商品詳細表示時：view_item + stock", "view_item", "高", "正常"),
    ("TC-041", "商品・カート", "カート追加成功時：add_to_cart（add_method=normal）", "add_to_cart", "高", "正常"),
    ("TC-042", "商品・カート", "add_to_cart：add_method=quick_order", "add_to_cart", "中", "正常"),
    ("TC-043", "商品・カート", "add_to_cart：add_method=order_history", "add_to_cart", "中", "正常"),
    ("TC-044", "商品・カート", "add_to_cart：add_method=quote_history", "add_to_cart", "中", "正常"),
    ("TC-045", "商品・カート", "カート削除時：remove_from_cart", "remove_from_cart", "高", "正常"),
    ("TC-046", "商品・カート", "お気に入り追加時：add_to_wishlist", "add_to_wishlist", "高", "正常"),
    ("TC-047", "商品・カート", "ecommerceイベント前にecommerce:nullをpush", "ecommerce clear", "高", "正常"),
    ("TC-048", "商品・カート", "カート追加失敗時はadd_to_cartを送信しない", "add_to_cart", "高", "異常"),
    # CHECKOUT
    ("TC-050", "購入フロー", "注文プロセス開始：begin_checkout（cart_type=normal）", "begin_checkout", "高", "正常"),
    ("TC-051", "購入フロー", "begin_checkout（cart_type=quote）", "begin_checkout", "高", "正常"),
    ("TC-052", "購入フロー", "request_item_countの値（あり/なし=0）", "begin_checkout", "中", "正常"),
    ("TC-060", "購入フロー", "add_shipping_infoとadd_payment_infoの同時送信", "add_shipping_info/add_payment_info", "高", "正常"),
    ("TC-061", "購入フロー", "payment_typeの値（4種）", "add_payment_info", "中", "正常"),
    ("TC-062", "購入フロー", "クーポン適用時：ecommerce.couponの送信", "add_shipping_info/add_payment_info", "中", "正常"),
    ("TC-063", "購入フロー", "クーポン未適用時：couponキーを出力しない", "add_shipping_info/add_payment_info", "中", "異常"),
    ("TC-064", "購入フロー", "配送・決済情報のサーバー受理失敗時はadd_shipping_info/add_payment_infoを送信しない", "add_shipping_info/add_payment_info", "高", "異常"),
    # PURCHASE
    ("TC-070", "注文確定", "承認申請時：approval_request（order_id）", "approval_request", "高", "正常"),
    ("TC-071", "注文確定", "注文確定時：purchase（承認なし）", "purchase", "高", "正常"),
    ("TC-072", "注文確定", "承認あり：承認完了時にpurchase送信", "purchase", "高", "正常"),
    ("TC-073", "注文確定", "決済エラー時はpurchaseを送信しない", "purchase", "高", "異常"),
    ("TC-074", "注文確定", "customer_type（new/returning/guest）", "purchase", "高", "正常"),
    ("TC-075", "注文確定", "order_id = transaction_id = approval_request.order_id", "purchase/approval_request", "高", "正常"),
    ("TC-076", "注文確定", "承認不要の注文でapproval_requestを送信しない", "approval_request", "高", "異常"),
    ("TC-077", "注文確定", "承認申請のみではpurchaseを送信しない", "purchase", "高", "異常"),
    # OTHER
    ("TC-080", "その他", "マイページ：mailmagazine_signup（all/text_only/none）", "mailmagazine_signup", "中", "正常"),
    ("TC-081", "その他", "一括注文フォーマットDL：bulk_order_download", "bulk_order_download", "高", "正常"),
    ("TC-082", "その他", "一括注文アップロード成功：bulk_order_upload + add_to_cart複数", "bulk_order_upload", "高", "正常"),
    ("TC-083", "その他", "一括注文アップロード失敗：bulk_order_uploadのみ（add_to_cartなし）", "bulk_order_upload", "高", "異常"),
    ("TC-084", "その他", "error_typeの複数値（カンマ区切り）", "bulk_order_upload", "中", "正常"),
    ("TC-085", "その他", "マイページ：メルマガ設定保存失敗時はmailmagazine_signupを送信しない", "mailmagazine_signup", "中", "異常"),
    # ITEMS
    ("TC-090", "items配列", "items共通フィールドの存在・型", "items", "高", "正常"),
    ("TC-091", "items配列", "item_variantなし時はキーごと省略", "items", "中", "正常"),
    ("TC-092", "items配列", "サプライヤー異なる商品は別要素として配列に含む", "items", "中", "正常"),
]

DETAILS = {
    "TC-001": {
        "前提": f"対象サイト（{SITE}）の任意ページにアクセスできること",
        "手順": "1. ブラウザで任意ページを開く\n2. DevTools Elementsで<head>内を確認",
        "コマンド": "document.querySelector('script[src*=\"googletagmanager.com/gtm.js\"]')",
        "期待": f"null以外。srcに「{GTM_ID}」が含まれる。",
    },
    "TC-002": {
        "前提": "同上",
        "手順": "1. Ctrl+U でページソースを表示\n2. <body> 直後に GTM noscript があるか目視確認\n※ Console の querySelector は使わない（JS有効時は noscript 内が DOM 化されないため）",
        "コマンド": "document.documentElement.innerHTML.includes('googletagmanager.com/ns.html?id=GTM-TGF9D8WX')",
        "期待": "true が返る。ソース上に <!-- Google Tag Manager (noscript) --> と iframe が存在する。",
    },
    "TC-003": {
        "前提": "ログイン済み状態",
        "手順": "1. Ctrl+U でページソースを表示\n2. <head> 内で user_id の push が GTM スニペットより前にあるか確認\n3. Console で dataLayer の順序も補助確認",
        "コマンド": "(() => { const uidIdx = dataLayer.findIndex(d => d.user_id !== undefined); const gtmIdx = dataLayer.findIndex(d => d['gtm.start'] !== undefined); return { user_id_index: uidIdx, gtm_start_index: gtmIdx, ok: uidIdx >= 0 && gtmIdx >= 0 && uidIdx < gtmIdx }; })()",
        "期待": "ソース上で user_id push が GTM より前。Console補助: { ok: true }",
    },
    "TC-004": {
        "前提": "任意ページを初回読み込み",
        "手順": "1. ページ読み込み直後にConsoleを開く",
        "コマンド": "Array.isArray(window.dataLayer)",
        "期待": "true。dataLayer は配列として存在する。",
    },
    "TC-005": {
        "前提": "任意ページ",
        "手順": "1. ページソースで gtm.js を検索\n2. Console で gtm.start の件数を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'gtm.js').length",
        "期待": "1（GTMスニペットが重複していないこと）。2以上はNG。",
    },
    "TC-010": {
        "前提": "取引先・担当者でログイン済み",
        "手順": "1. ページを開き直す（フルリロード）\n2. Consoleで確認",
        "コマンド": "dataLayer.find(d => d.user_id)",
        "期待": "{ user_id, area_id, staff_id, member_r } が存在。member_r は数値。",
    },
    "TC-011": {
        "前提": "未ログイン（ゲスト）状態",
        "手順": "1. シークレットウィンドウで開く、またはログアウト後に Ctrl+Shift+R\n2. Consoleで確認\n※ Turbo遷移直後のみ確認しない（旧セッションが残るため）",
        "コマンド": "dataLayer.filter(d => 'user_id' in d)",
        "期待": "[]（空配列）。user_id キーを持つオブジェクトが存在しない。",
    },
    "TC-012": {
        "前提": "会員ランクが異なるテストアカウント",
        "手順": "1. 標準(0)/シルバー(1)/ゴールド(2)/その他(3)でそれぞれログイン\n2. member_r を確認",
        "コマンド": "dataLayer.find(d => d.member_r !== undefined)?.member_r",
        "期待": "0 / 1 / 2 / 3（数値型）。",
    },
    "TC-013": {
        "前提": "ログイン済み",
        "手順": "1. dataLayer.length を記録\n2. F5 でリロード\n3. 再度 length と [0] を確認",
        "コマンド": "({ length: dataLayer.length, first: dataLayer[0] })",
        "期待": "リロード後も初期件数（例:4）に戻る。[0] に user_id が再出力される。蓄積しない。",
    },
    "TC-014": {
        "前提": "ログイン済み。ログイン操作直後ではない通常ページ",
        "手順": "1. ログイン済み状態でトップ等をフルリロード\n2. login イベントの有無を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'login')",
        "期待": "[]。通常ページ表示だけでは login は送信されない。",
    },
    "TC-020": {
        "前提": "未ログイン。有効なテストアカウント",
        "手順": "1. 正しいID/PWでログイン\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'login')",
        "期待": "1件以上。{ event: 'login' } を含む。\n※ gtm.uniqueEventId は GTM が自動付与するためあってもOK。",
    },
    "TC-021": {
        "前提": "未ログイン",
        "手順": "1. 誤ったID/PWでログイン試行\n2. 操作前後で件数を比較",
        "コマンド": "(() => { const n = dataLayer.filter(d => d.event === 'login').length; return { count: n }; })()",
        "期待": "count: 0。login イベントは追加されない。",
    },
    "TC-022": {
        "前提": "会員登録フォームにアクセス可能",
        "手順": "1. 必須項目を正しく入力して送信\n2. サーバー側で登録成功後に確認",
        "コマンド": "dataLayer.filter(d => d.event === 'sign_up')",
        "期待": "[{ event: 'sign_up' }] が1件以上。",
    },
    "TC-023": {
        "前提": "会員登録フォームにアクセス可能",
        "手順": "1. 必須項目未入力・形式エラー・重複メール等で送信\n2. サーバー側で登録失敗後に確認",
        "コマンド": "dataLayer.filter(d => d.event === 'sign_up')",
        "期待": "[]。登録失敗時は sign_up を送信しない。",
    },
    "TC-024": {
        "前提": "会員登録フォーム。メルマガ「全て受け取る」を選択可能",
        "手順": "1. メルマガ受信設定を「全て受け取る」にして登録成功\n2. sign_up と mailmagazine_signup を確認",
        "コマンド": "dataLayer.filter(d => ['sign_up','mailmagazine_signup'].includes(d.event)).map(d => ({ event: d.event, subscription_status: d.subscription_status }))",
        "期待": "sign_up と mailmagazine_signup の両方が存在。mailmagazine_signup は subscription_status: 'all'。",
    },
    "TC-025": {
        "前提": "会員登録フォーム。メルマガ「テキストのみ」を選択可能",
        "手順": "1. メルマガ受信設定を「テキストのみ」にして登録成功\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'mailmagazine_signup').pop()?.subscription_status",
        "期待": "'text_only'",
    },
    "TC-026": {
        "前提": "会員登録フォーム。メルマガ「受け取らない」を選択可能",
        "手順": "1. メルマガ受信設定を「受け取らない」にして登録成功\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'mailmagazine_signup').pop()?.subscription_status",
        "期待": "'none'",
    },
    "TC-027": {
        "前提": "会員登録フォーム",
        "手順": "1. 登録失敗（バリデーションエラー等）を発生させる\n2. sign_up / mailmagazine_signup を確認",
        "コマンド": "dataLayer.filter(d => ['sign_up','mailmagazine_signup'].includes(d.event))",
        "期待": "[]。登録失敗時は sign_up も mailmagazine_signup も送信しない。",
    },
    "TC-030": {
        "前提": "見積書DL可能な状態",
        "手順": "1. 見積書DLリンクをクリック\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'quote_download')",
        "期待": "[{ event: 'quote_download' }]",
    },
    "TC-031": {
        "前提": "ログイン済み。新規見積依頼可能",
        "手順": "1. 見積依頼フォームを送信して成功\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'generate_lead').map(d => ({ lead_source: d.lead_source, items: d.ecommerce?.items?.length }))",
        "期待": "lead_source: 'quote'、items 配列あり。",
    },
    "TC-032": {
        "前提": "問い合わせフォームにアクセス可能",
        "手順": "1. 問い合わせフォームを送信して成功\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'generate_lead').pop()",
        "期待": "{ event: 'generate_lead', lead_source: 'inquiry' }。items は含まれない。",
    },
    "TC-033": {
        "前提": "再見積可能な既存見積あり",
        "手順": "1. 操作前の generate_lead 件数を記録\n2. 再見積を実行\n3. 件数を比較",
        "コマンド": "dataLayer.filter(d => d.event === 'generate_lead').length",
        "期待": "再見積操作後も件数が増えない。",
    },
    "TC-034": {
        "前提": "見積依頼フォームにアクセス可能",
        "手順": "1. 必須項目未入力等で送信しバリデーションエラーを発生\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'generate_lead')",
        "期待": "[]。バリデーションエラー時は generate_lead を送信しない。",
    },
    "TC-035": {
        "前提": "問い合わせフォームにアクセス可能",
        "手順": "1. 必須項目未入力等で送信しエラーを発生\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'generate_lead')",
        "期待": "[]。送信失敗時は generate_lead を送信しない。",
    },
    "TC-040": {
        "前提": "商品詳細ページにアクセス可能",
        "手順": "1. 商品詳細を表示\n2. view_item を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'view_item').pop()?.ecommerce?.items?.[0]",
        "期待": "item_id, item_name, item_brand, item_category1〜5, price, quantity, stock（数値）が含まれる。",
    },
    "TC-041": {
        "前提": "ログイン済み。カート追加可能",
        "手順": "1. 通常操作でカートに追加\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').pop()",
        "期待": "{ event:'add_to_cart', add_method:'normal', ecommerce:{ value, currency:'JPY', items } }",
    },
    "TC-042": {
        "前提": "クイックオーダー利用可能",
        "手順": "1. クイックオーダー経由でカート追加\n2. add_method を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').pop()?.add_method",
        "期待": "'quick_order'",
    },
    "TC-043": {
        "前提": "購入履歴からの再注文が可能",
        "手順": "1. 購入履歴からカート追加\n2. add_method を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').pop()?.add_method",
        "期待": "'order_history'",
    },
    "TC-044": {
        "前提": "見積履歴からのカート追加が可能",
        "手順": "1. 見積履歴からカート追加\n2. add_method を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').pop()?.add_method",
        "期待": "'quote_history'",
    },
    "TC-045": {
        "前提": "カートに商品あり",
        "手順": "1. カートから商品を削除\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'remove_from_cart').pop()",
        "期待": "{ event:'remove_from_cart', ecommerce:{ value, currency:'JPY', items } }",
    },
    "TC-046": {
        "前提": "商品詳細ページにアクセス可能",
        "手順": "1. お気に入りボタンをクリック\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_wishlist').pop()",
        "期待": "{ event:'add_to_wishlist', ecommerce:{ currency:'JPY', items } }",
    },
    "TC-047": {
        "前提": "ecommerce系イベントを発火させる操作",
        "手順": "1. view_item または add_to_cart を実行\n2. 直前の push を確認",
        "コマンド": "(() => { const i = dataLayer.findIndex(d => d.event === 'view_item' || d.event === 'add_to_cart'); return i >= 1 ? dataLayer[i-1] : null; })()",
        "期待": "{ ecommerce: null }",
    },
    "TC-048": {
        "前提": "カート追加失敗を再現できる（在庫切れ・数量エラー等）",
        "手順": "1. 操作前の add_to_cart 件数を記録\n2. カート追加失敗を発生\n3. 件数を比較",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').length",
        "期待": "失敗操作後も件数が増えない。",
    },
    "TC-050": {
        "前提": "通常カートに商品あり",
        "手順": "1. 注文プロセスへ進む\n2. begin_checkout を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'begin_checkout').pop()",
        "期待": "cart_type:'normal', request_item_count: 0, ecommerce あり。",
    },
    "TC-051": {
        "前提": "見積カートに商品あり",
        "手順": "1. 見積カートから注文プロセスへ進む\n2. cart_type を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'begin_checkout').pop()?.cart_type",
        "期待": "'quote'",
    },
    "TC-052": {
        "前提": "リクエスト注文あり/なしの両ケース",
        "手順": "1-A. 通常商品のみで開始\n1-B. リクエスト注文含めて開始",
        "コマンド": "dataLayer.filter(d => d.event === 'begin_checkout').pop()?.request_item_count",
        "期待": "通常のみ: 0。リクエスト注文含む: 正の整数。",
    },
    "TC-060": {
        "前提": "注文プロセス中",
        "手順": "1. 配送・決済情報を入力してサーバー受理成功\n2. 両イベントを確認",
        "コマンド": "dataLayer.filter(d => ['add_shipping_info','add_payment_info'].includes(d.event)).map(d => d.event)",
        "期待": "['add_shipping_info', 'add_payment_info'] の両方。",
    },
    "TC-061": {
        "前提": "各決済方法で注文可能",
        "手順": "1. 各決済方法で送信\n2. payment_type を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_payment_info').pop()?.payment_type",
        "期待": "on_credit / cash_on_delivery / bank_transfer / full_point のいずれか。",
    },
    "TC-062": {
        "前提": "クーポン適用可能な注文",
        "手順": "1. クーポン適用後に配送・決済情報を送信\n2. coupon を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_shipping_info').pop()?.ecommerce?.coupon",
        "期待": "クーポンコード文字列が返る。",
    },
    "TC-063": {
        "前提": "クーポン未適用の注文",
        "手順": "1. クーポンなしで配送・決済情報を送信\n2. coupon キーの有無を確認",
        "コマンド": "'coupon' in (dataLayer.filter(d => d.event === 'add_shipping_info').pop()?.ecommerce || {})",
        "期待": "false（coupon キーごと省略）。",
    },
    "TC-064": {
        "前提": "配送・決済情報の送信失敗を再現できる",
        "手順": "1. 操作前の件数を記録\n2. サーバー受理失敗を発生\n3. 件数を比較",
        "コマンド": "dataLayer.filter(d => ['add_shipping_info','add_payment_info'].includes(d.event)).length",
        "期待": "失敗操作後も件数が増えない。",
    },
    "TC-070": {
        "前提": "承認フロー対象の注文が可能",
        "手順": "1. 承認申請ボタンを押下\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'approval_request').pop()",
        "期待": "{ event:'approval_request', order_id:'T_...' }",
    },
    "TC-071": {
        "前提": "承認不要の注文が可能",
        "手順": "1. 注文確定して完了\n2. purchase を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'purchase').pop()",
        "期待": "order_id, customer_type, point_amount, cart_type, ecommerce（transaction_id, value, tax, shipping 等）あり。",
    },
    "TC-072": {
        "前提": "承認フロー対象。承認者アカウントあり",
        "手順": "1. 承認申請\n2. 承認者が承認完了\n3. purchase の発火タイミングを確認",
        "コマンド": "dataLayer.filter(d => d.event === 'purchase')",
        "期待": "承認完了時に purchase が送信される（申請時ではない）。",
    },
    "TC-073": {
        "前提": "決済エラーを再現できる",
        "手順": "1. 操作前の purchase 件数を記録\n2. 決済エラーで注文確定を試行\n3. 件数を比較",
        "コマンド": "dataLayer.filter(d => d.event === 'purchase').length",
        "期待": "件数が増えない。",
    },
    "TC-074": {
        "前提": "初回/リピート/ゲストの各ケース",
        "手順": "1. 各ケースで注文完了\n2. customer_type を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'purchase').pop()?.customer_type",
        "期待": "new / returning / guest のいずれか。",
    },
    "TC-075": {
        "前提": "承認フローありの注文完了",
        "手順": "1. 承認申請→承認→購入完了\n2. IDを比較",
        "コマンド": "(() => { const ar = dataLayer.find(d=>d.event==='approval_request'); const pu = dataLayer.find(d=>d.event==='purchase'); return {approval: ar?.order_id, transaction: pu?.ecommerce?.transaction_id, order: pu?.order_id}; })()",
        "期待": "3つの値がすべて同一。",
    },
    "TC-076": {
        "前提": "承認不要の注文が可能",
        "手順": "1. 承認不要で注文確定\n2. approval_request の有無を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'approval_request')",
        "期待": "[]。承認不要注文では approval_request を送信しない。",
    },
    "TC-077": {
        "前提": "承認フロー対象の注文",
        "手順": "1. 承認申請のみ実行（承認完了前）\n2. purchase の有無を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'purchase')",
        "期待": "[]。承認申請時点では purchase を送信しない。",
    },
    "TC-080": {
        "前提": "ログイン済み。マイページでメルマガ設定変更可能",
        "手順": "1. マイページで受信設定を all / text_only / none にそれぞれ変更して保存成功\n2. 各 subscription_status を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'mailmagazine_signup').map(d => d.subscription_status)",
        "期待": "変更ごとに 'all' / 'text_only' / 'none' が送信される。",
    },
    "TC-081": {
        "前提": "一括注文画面にアクセス可能",
        "手順": "1. 一括注文CSVフォーマットをDL\n2. Consoleで確認",
        "コマンド": "dataLayer.filter(d => d.event === 'bulk_order_download')",
        "期待": "[{ event: 'bulk_order_download' }]",
    },
    "TC-082": {
        "前提": "正しいCSVファイル",
        "手順": "1. CSVアップロード成功\n2. upload と add_to_cart を確認",
        "コマンド": "({ upload: dataLayer.filter(d => d.event === 'bulk_order_upload'), carts: dataLayer.filter(d => d.event === 'add_to_cart' && d.add_method === 'bulk_order') })",
        "期待": "upload_status:'success'。全商品分の add_to_cart（add_method:'bulk_order'）あり。",
    },
    "TC-083": {
        "前提": "エラーになるCSV",
        "手順": "1. エラーCSVをアップロード\n2. upload と add_to_cart を確認",
        "コマンド": "({ upload: dataLayer.filter(d => d.event === 'bulk_order_upload'), carts: dataLayer.filter(d => d.event === 'add_to_cart') })",
        "期待": "upload_status:'error'。add_to_cart は増えない。",
    },
    "TC-084": {
        "前提": "複数エラーが同時に発生するCSV",
        "手順": "1. 複合エラーCSVをアップロード\n2. error_type を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'bulk_order_upload').pop()?.error_type",
        "期待": "カンマ区切りで複数エラー（例: 'quantity_invalid_multiple,insufficient_stock'）。",
    },
    "TC-085": {
        "前提": "マイページでメルマガ設定変更可能",
        "手順": "1. 操作前の件数を記録\n2. 保存失敗（バリデーションエラー等）を発生\n3. 件数を比較",
        "コマンド": "dataLayer.filter(d => d.event === 'mailmagazine_signup').length",
        "期待": "失敗操作後も件数が増えない。",
    },
    "TC-090": {
        "前提": "ecommerce.items を含むイベント発火済み",
        "手順": "1. view_item または add_to_cart を発火\n2. items[0] を確認",
        "コマンド": "Object.keys(dataLayer.find(d => d.ecommerce?.items).ecommerce.items[0])",
        "期待": "item_id, item_name, item_brand, item_category1〜5, price, quantity が含まれる。price/quantity は数値。",
    },
    "TC-091": {
        "前提": "バリエーションなし商品",
        "手順": "1. バリエーションなし商品の詳細を表示\n2. item_variant の有無を確認",
        "コマンド": "'item_variant' in (dataLayer.find(d => d.event === 'view_item')?.ecommerce?.items?.[0] || {})",
        "期待": "false。",
    },
    "TC-092": {
        "前提": "異なるサプライヤーの商品を同時にカート追加可能",
        "手順": "1. サプライヤーAとBの商品をカートに追加\n2. items 配列を確認",
        "コマンド": "dataLayer.filter(d => d.event === 'add_to_cart').pop()?.ecommerce?.items?.map(i => i.item_brand)",
        "期待": "異なるサプライヤー名が別要素として含まれる。",
    },
}

CATEGORY_SHEETS = {
    "GTM基本設定": "TC-GTM",
    "共通ページ計測": "TC-COMMON",
    "認証": "TC-AUTH",
    "リード": "TC-LEAD",
    "商品・カート": "TC-ECOM",
    "購入フロー": "TC-CHECKOUT",
    "注文確定": "TC-PURCHASE",
    "その他": "TC-OTHER",
    "items配列": "TC-ITEMS",
}


def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_matrix_sheet(wb):
    ws = wb.active
    ws.title = "テストマトリクス"
    headers = ["テストID", "カテゴリ", "種別", "テスト項目", "対象イベント/機能", "優先度", "実施結果", "実施日", "実施者", "備考"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [10, 14, 8, 40, 28, 8, 10, 12, 10, 20])

    for r, tc in enumerate(TEST_CASES, 2):
        tc_id, cat, name, target, pri, kind = tc
        row_vals = [tc_id, cat, kind, name, target, pri, "", "", "", ""]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORMAL
            cell.border = BORDER
            cell.alignment = WRAP

    lr = len(TEST_CASES) + 3
    ws.cell(row=lr, column=1, value="【凡例】").font = Font(bold=True, size=10)
    ws.cell(row=lr + 1, column=1, value="種別: 正常=成功時に送信されること / 異常=失敗時・非該当時に送信されないこと")
    ws.cell(row=lr + 2, column=1, value="実施結果: OK / NG / 未実施 / 保留")
    ws.cell(row=lr + 3, column=1, value=f"対象サイト: {SITE}")
    ws.cell(row=lr + 4, column=1, value=f"GTMコンテナID: {GTM_ID}")


def create_guide_sheet(wb):
    ws = wb.create_sheet("テスト実施手順（共通）", 1)
    rows = [
        ["Primo dataLayer 手動テスト実施手順"],
        [""],
        ["■ 事前準備"],
        ["1. Chrome（推奨）で対象サイトにアクセス"],
        ["2. F12 → Consoleタブを開く"],
        ["3. 異常系テストは操作前後でイベント件数を比較する"],
        ["4. 未ログイン確認はシークレットウィンドウまたは Ctrl+Shift+R を使用"],
        [""],
        ["■ 基本確認コマンド"],
        ["コマンド", "用途", "期待される出力例"],
        ["dataLayer.filter(d => d.event === 'イベント名')", "特定イベントを抽出", "該当イベントの配列"],
        ["dataLayer.filter(d => d.event === 'イベント名').length", "件数比較（異常系）", "操作前後で増えないこと"],
        ["dataLayer.find(d => d.user_id)", "user_id付きpush", "{user_id, area_id, staff_id, member_r}"],
        ["dataLayer.filter(d => 'user_id' in d)", "未ログイン確認", "[]"],
        [""],
        ["■ 注意事項"],
        ["・gtm.uniqueEventId は GTM が自動付与。実装の余計なキーとはみなさない"],
        ["・noscript（TC-002）はページソース（Ctrl+U）で確認。querySelector は使わない"],
        ["・TC-003 はページソースの記述順序 + dataLayer 順序で確認"],
        ["・Turbo遷移直後は旧 dataLayer が残ることがある。フルリロードで判定"],
        ["・異常系は『イベントが送信されない』ことを確認する"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=14, color="1A3A5C")
            elif row and str(row[0]).startswith("■"):
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = NORMAL
            cell.alignment = WRAP
    set_col_widths(ws, [70, 30, 40])
    ws.merge_cells("A1:C1")


def create_detail_sheet(wb, sheet_name, category):
    ws = wb.create_sheet(sheet_name)
    cases = [tc for tc in TEST_CASES if tc[1] == category]
    if not cases:
        return
    headers = ["テストID", "種別", "テスト項目", "前提条件", "操作手順", "確認コマンド（Console）", "期待結果（コマンド出力）", "実施結果", "備考"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [10, 8, 26, 22, 28, 38, 42, 10, 16])

    row = 2
    for tc_id, _, name, _, priority, kind in cases:
        d = DETAILS[tc_id]
        values = [tc_id, kind, f"【{priority}】{name}", d["前提"], d["手順"], d["コマンド"], d["期待"], "", ""]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL
            cell.border = BORDER
            cell.alignment = WRAP
        ws.row_dimensions[row].height = 130
        row += 1
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    create_matrix_sheet(wb)
    create_guide_sheet(wb)
    for category, sheet_name in CATEGORY_SHEETS.items():
        create_detail_sheet(wb, sheet_name, category)
    out = "/workspace/Primo_dataLayer実装単体テスト.xlsx"
    wb.save(out)
    normal = sum(1 for t in TEST_CASES if t[5] == "正常")
    abnormal = sum(1 for t in TEST_CASES if t[5] == "異常")
    print(f"Saved: {out}")
    print(f"Total: {len(TEST_CASES)} (正常:{normal} / 異常:{abnormal})")


if __name__ == "__main__":
    main()
