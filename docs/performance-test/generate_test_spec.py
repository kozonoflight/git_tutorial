#!/usr/bin/env python3
"""Generate performance/load test specification Excel for Flight Payment Center."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "docs/performance-test/フライト決済センター_性能負荷テスト仕様書.xlsx"

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
MANUAL_FILL = PatternFill("solid", fgColor="E2EFDA")
TOOL_FILL = PatternFill("solid", fgColor="FCE4D6")
K6_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

RAMP_COLS = [f"同時{n}人" for n in range(1, 11)]
LOAD_TOOL = "k6"

# All front-site screens (from requirements)
FRONT_SCREENS_DISPLAY = [
    "クレジットカード一覧",
    "クレジットカード登録画面（フライト決済センター）",
    "クレジットカード登録画面からの戻り",
    "クレジットカード削除",
    "クレカ使用での購入",
    "クレカ使用での見積購入",
    "お支払方法選択画面",
    "見積購入でのお支払方法選択画面",
]

FRONT_SCREENS_CONCURRENT = [
    "クレジットカード登録",
    "クレジットカード削除",
    "クレカ使用での購入",
    "クレカ使用での見積購入",
]

ADMIN_SCREENS = [
    "出荷処理",
    "注文変更処理",
    "注文キャンセル処理",
]

LARGE_DATA_PRECONDITION = (
    "1注文あたり9サプライヤー、1サプライヤーあたり10商品（計90商品）の注文データを用意。"
    "k6で同時ユーザー数を1人→2人→3人…と段階的に増加させて負荷をかける。"
)

BATCH_STUB_NOTE = (
    "テストサイトのためフライト決済センター本体に負荷をかけられない。"
    "100注文分のデータを用意し、フライト決済センター呼び出しはスタブで代替して実施する。"
)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_area(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def empty_ramp_tail():
    return [""] * (len(RAMP_COLS) + 3)  # ramp cols + 総合判定 + 実施日 + 備考


def build_overview(wb):
    ws = wb.active
    ws.title = "概要"
    rows = [
        ["フライト決済センター 性能・負荷テスト仕様書", ""],
        ["", ""],
        ["文書バージョン", "1.1"],
        ["作成日", "2026-07-08"],
        ["対象システム", "フライト決済センター（フロントサイト / 管理サイト）"],
        ["負荷テストツール", "k6"],
        ["", ""],
        ["テスト目的", "新規作成画面・処理の性能および負荷耐性を確認する"],
        ["", ""],
        ["テスト区分", "内容", "実施方法", "合格基準"],
        [
            "1. 画面表示の秒数", "画面描画までの処理時間",
            "手動（同時1人→10人を段階的に実施）", "3秒以内",
        ],
        [
            "2. 同時操作の正常終了", "複数ユーザー同時操作時の正常完了",
            "手動（同時1人→10人を段階的に実施）", "DB・フライト処理が正常完了",
        ],
        [
            "3. 大量データの正常終了", "90商品/注文（9サプライヤー×10商品）条件下の処理完了",
            f"k6（同時1人→10人を段階的に実施）", "DB・フライト処理が正常完了",
        ],
        [
            "4. バッチ処理の大量データ処理", "100注文規模のバッチ正常完了",
            "データ投入 + フライト決済センタースタブ", "サイクル内完了・多重実行なし",
        ],
        ["", ""],
        ["段階的負荷テスト（人数増加）", ""],
        [
            "対象",
            "テスト項目1（画面表示の秒数）、テスト項目2（同時操作の正常終了）、"
            "テスト項目3（大量データの正常終了）",
        ],
        [
            "実施方法",
            "同時操作人数を1人、2人、3人…と段階的に増やし、各段階での秒数・正常終了を確認する。"
            "最大10人まで実施する。",
        ],
        [
            "テスト項目3のデータ条件",
            "1注文に9サプライヤー、1サプライヤーに10商品ずつ（計90商品）を注文商品とする。",
        ],
        ["", ""],
        ["バッチテストの補足", ""],
        [
            "スタブ利用",
            "テストサイトのためフライト決済センターに直接負荷をかけられない。"
            "100注文に対してバッチを実施し、決済センター呼び出しはスタブで代替する。",
        ],
        ["", ""],
        ["確認対象テーブル", "orders（注文）, fcmp_transaction_histories（クレジット履歴） ほか"],
        ["", ""],
        ["対象画面・処理（フロントサイト）", ""],
    ]
    for screen in FRONT_SCREENS_DISPLAY:
        rows.append(["", screen])
    rows += [["", ""], ["対象画面・処理（管理サイト）", ""]]
    for screen in ADMIN_SCREENS:
        rows.append(["", screen])
    rows += [
        ["", ""],
        ["対象バッチ", ""],
        ["", "再オーソリバッチ"],
        ["", "障害取消実行バッチ（5分毎実行）"],
        ["", "定期購入バッチ（契約サイト・100件）"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = CENTER
    for r in (9, 15, 19):
        style_header_row(ws, r, 4)
    set_col_widths(ws, [28, 48, 28, 28])
    ws.row_dimensions[1].height = 30


def build_ramp_sheet(wb):
    """Dedicated sheet explaining ramp-up test matrix."""
    ws = wb.create_sheet("段階的負荷テスト方針")
    rows = [
        ["段階的負荷テスト方針（同時人数の増加）", ""],
        ["", ""],
        ["項目", "内容"],
        ["対象テスト区分", "1.画面表示の秒数 / 2.同時操作の正常終了 / 3.大量データの正常終了"],
        ["負荷ツール", "テスト項目1・2: 手動（人数を段階的に増加） / テスト項目3: k6"],
        ["同時人数", "1人 → 2人 → 3人 → … → 10人（各段階で計測・判定）"],
        ["", ""],
        ["テスト区分", "各段階での確認内容", "記録欄"],
        ["1. 画面表示の秒数", "各画面の表示秒数（s）が3秒以内か", "詳細シートの「同時N人」列に秒数を記録"],
        ["2. 同時操作の正常終了", "fcmp_transaction_histories・フライト処理が正常完了か", "詳細シートの「同時N人」列にOK/NGを記録"],
        [
            "3. 大量データの正常終了",
            "90商品/注文（9サプライヤー×10商品）でk6負荷時に正常完了か",
            "詳細シートの「同時N人」列にOK/NGを記録",
        ],
        ["", ""],
        ["大量データの注文条件", ""],
        ["サプライヤー数", "9（1注文あたり）"],
        ["1サプライヤーあたりの商品数", "10"],
        ["1注文あたりの合計商品数", "90"],
        ["負荷のかけ方", "k6で同時ユーザー数を1→2→3…と段階的に増加"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(bold=True, size=14)
    style_header_row(ws, 3, 2)
    style_header_row(ws, 8, 3)
    set_col_widths(ws, [28, 50, 30])


def build_summary(wb):
    ws = wb.create_sheet("テストケース一覧")
    headers = [
        "テストID", "テスト区分", "実施方法", "対象画面/処理", "サイト",
        "確認ポイント", "段階的負荷", "優先度", "結果", "実施者", "実施日", "備考",
    ]
    cases = []
    pid = 1

    for screen in FRONT_SCREENS_DISPLAY:
        cases.append((
            f"PT-{pid:03d}", "1.画面表示", "手動(1→10人)", screen, "フロント",
            "3秒以内表示", "1〜10人", "高", "", "", "", "",
        ))
        pid += 1

    for screen in FRONT_SCREENS_CONCURRENT:
        cases.append((
            f"PT-{pid:03d}", "2.同時操作", "手動(1→10人)", screen, "フロント",
            "fcmp_transaction_histories正常", "1〜10人", "高", "", "", "", "",
        ))
        pid += 1

    for screen in ADMIN_SCREENS:
        cases.append((
            f"PT-{pid:03d}", "2.同時操作", "手動(1→10人)", screen, "管理",
            "処理・決済正常完了", "1〜10人", "高", "", "", "", "他管理処理と同時実施可",
        ))
        pid += 1

    cases.append((
        f"PT-{pid:03d}", "2.同時操作", "手動(1→10人)",
        "出荷+注文変更+キャンセル同時", "管理",
        "各処理が干渉せず正常完了", "1〜10人", "高", "", "", "", "組合せ同時操作",
    ))
    pid += 1

    large_screens = [
        "お支払方法選択画面",
        "見積購入でのお支払方法選択画面",
        "出荷処理",
        "注文変更処理",
        "注文キャンセル処理",
    ]
    for screen in large_screens:
        site = "フロント" if "画面" in screen else "管理"
        cases.append((
            f"PT-{pid:03d}", "3.大量データ", f"k6(1→10人)", screen, site,
            "90商品/注文で正常完了", "1〜10人", "高", "", "", "",
            "9サプライヤー×10商品=90商品",
        ))
        pid += 1

    for batch, site, note in [
        ("再オーソリバッチ", "バッチ", "100注文・スタブ利用"),
        ("障害取消実行バッチ（5分毎）", "バッチ", "100注文・スタブ利用・5分超過時確認"),
        ("定期購入バッチ", "契約サイト", "100件・契約サイトで実施"),
    ]:
        cases.append((
            f"PT-{pid:03d}", "4.バッチ", "スタブ+データ投入", batch, site,
            "100注文処理・多重実行なし", "—", "高", "", "", "", note,
        ))
        pid += 1

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    for r, case in enumerate(cases, 2):
        for c, val in enumerate(case, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 3:
                if "手動" in str(val):
                    cell.fill = MANUAL_FILL
                elif "k6" in str(val):
                    cell.fill = K6_FILL
                else:
                    cell.fill = TOOL_FILL
    style_data_area(ws, 2, len(cases) + 1, len(headers))
    set_col_widths(ws, [10, 14, 14, 34, 10, 26, 10, 8, 8, 10, 12, 24])
    ws.freeze_panes = "A2"


def build_detail_sheet(wb, title, section_no, method, pass_criteria, cases, value_hint=""):
    ws = wb.create_sheet(title)
    base_headers = [
        "テストID", "対象画面/処理", "サイト", "前提条件", "テスト手順",
        "期待結果", "合格基準", "実施方法",
    ]
    tail_headers = ["総合判定", "実施日", "備考"]
    headers = base_headers + RAMP_COLS + tail_headers

    ws.cell(row=1, column=1, value=f"テスト項目{section_no}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    ws.cell(row=2, column=1, value="実施方法")
    ws.cell(row=2, column=2, value=method)
    ws.cell(row=3, column=1, value="合格基準")
    ws.cell(row=3, column=2, value=pass_criteria)
    if value_hint:
        ws.cell(row=4, column=1, value="記録方法")
        ws.cell(row=4, column=2, value=value_hint)
        merge_end = 4
    else:
        merge_end = 3
    ws.merge_cells(start_row=2, start_column=2, end_row=merge_end, end_column=len(headers))
    ws.merge_cells(start_row=3, start_column=2, end_row=merge_end, end_column=len(headers))
    if value_hint:
        ws.merge_cells(start_row=4, start_column=2, end_row=merge_end, end_column=len(headers))

    header_row = merge_end + 2
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    data_start = header_row + 1
    for r, case in enumerate(cases, data_start):
        for c, val in enumerate(case, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_area(ws, data_start, data_start + len(cases) - 1, len(headers))
    set_col_widths(ws, [10, 30, 8, 28, 38, 28, 14, 12] + [9] * 10 + [10, 12, 20])
    ws.freeze_panes = f"A{data_start}"


def _display_steps(screen):
    if "購入" in screen and "見積" not in screen:
        return (
            "1. クレカ使用での購入フローを開始\n"
            "2. 購入確定/決済画面の描画完了までの秒数を計測（s）\n"
            "3. 同時人数を1→2→3…と増やし、各段階で計測"
        )
    if "見積購入" in screen:
        return (
            "1. クレカ使用での見積購入フローを開始\n"
            "2. 見積購入確定/決済画面の描画完了までの秒数を計測（s）\n"
            "3. 同時人数を1→2→3…と増やし、各段階で計測"
        )
    return (
        f"1. {screen}へ遷移する操作を実施\n"
        "2. 画面描画完了までの秒数を計測（s）\n"
        "3. 同時人数を1→2→3…と増やし、各段階で計測"
    )


def build_pt1(wb):
    cases = []
    for i, screen in enumerate(FRONT_SCREENS_DISPLAY, 1):
        pre = "テスト用アカウントでログイン済み。"
        if screen == "クレジットカード一覧":
            pre += "クレジットカードが1件以上登録済み。"
        elif screen == "クレジットカード削除":
            pre += "削除対象のクレジットカードが登録済み。"
        elif "購入" in screen:
            pre += "購入/見積購入可能な商品・クレジットカードを用意。"
        elif "支払方法" in screen:
            pre += "商品がカートに入っている。"
        cases.append((
            f"PT-{i:03d}", screen, "フロント", pre,
            _display_steps(screen),
            f"{screen}がエラーなく表示される", "3秒以内", "手動(1→10人)",
            *empty_ramp_tail(),
        ))
    build_detail_sheet(
        wb, "1_画面表示の秒数", 1,
        "手動テスト。同時操作人数を1人→2人→3人…→10人と段階的に増やし、各画面の表示秒数を計測する。",
        "ボタンクリックから画面描画完了まで3秒以内（単位: s）",
        cases,
        value_hint="「同時N人」列に各段階の表示秒数（s）を記録",
    )


def build_pt2(wb):
    base_id = len(FRONT_SCREENS_DISPLAY) + 1
    cases = []
    idx = 0

    concurrent_defs = [
        ("クレジットカード登録", "フロント",
         "テスト用アカウントを人数分用意。",
         "1. 指定人数が同時にクレジットカード登録を実行\n"
         "2. フライト決済センターで登録完了\n3. fcmp_transaction_historiesを確認\n"
         "4. 人数を1→2→3…と増やして繰り返す"),
        ("クレジットカード削除", "フロント",
         "各ユーザーに削除対象カードを用意。",
         "1. 指定人数が同時にクレジットカード削除を実行\n2. fcmp_transaction_historiesを確認\n"
         "3. 人数を1→2→3…と増やして繰り返す"),
        ("クレカ使用での購入", "フロント",
         "各ユーザーに購入可能な商品・カードを用意。",
         "1. 指定人数が同時にクレカ購入を実行\n2. orders, fcmp_transaction_historiesを確認\n"
         "3. フライト決済センター処理画面を確認\n4. 人数を1→2→3…と増やして繰り返す"),
        ("クレカ使用での見積購入", "フロント",
         "各ユーザーに見積購入可能な条件を用意。",
         "1. 指定人数が同時に見積購入（クレカ）を実行\n2. 関連テーブル・フライト処理を確認\n"
         "3. 人数を1→2→3…と増やして繰り返す"),
        ("出荷処理", "管理",
         "出荷対象の注文を各ユーザーに割当。",
         "1. 指定人数が同時に出荷処理を実行\n2. orders, fcmp_transaction_historiesを確認\n"
         "3. 人数を1→2→3…と増やして繰り返す"),
        ("注文変更処理", "管理",
         "変更対象の注文を各ユーザーに割当。",
         "1. 指定人数が同時に注文変更を実行\n2. 関連テーブル・フライト処理を確認\n"
         "3. 人数を1→2→3…と増やして繰り返す"),
        ("注文キャンセル処理", "管理",
         "キャンセル対象の注文を各ユーザーに割当。",
         "1. 指定人数が同時に注文キャンセルを実行\n2. 関連テーブル・フライト処理を確認\n"
         "3. 人数を1→2→3…と増やして繰り返す"),
    ]
    for screen, site, pre, steps in concurrent_defs:
        cases.append((
            f"PT-{base_id + idx:03d}", screen, site, pre, steps,
            "全操作が正常完了。DB・フライト処理に異常なし", "全件正常完了", "手動(1→10人)",
            *empty_ramp_tail(),
        ))
        idx += 1

    combo_tail = [""] * 12 + ["出荷と注文処理の同時挙動を重点確認"]
    cases.append((
        f"PT-{base_id + idx:03d}", "出荷+注文変更+キャンセル同時", "管理",
        "出荷・変更・キャンセル対象注文を混在させて用意。",
        "1. ユーザーが役割分担し、出荷・注文変更・キャンセルを同時実行\n"
        "2. 各処理の完了状態とDBを確認\n3. 人数を1→2→3…と増やして繰り返す",
        "処理が相互干渉せず全件正常完了", "全件正常完了", "手動(1→10人)",
        *combo_tail,
    ))

    build_detail_sheet(
        wb, "2_同時操作の正常終了", 2,
        "手動テスト。同時操作人数を1人→2人→3人…→10人と段階的に増やし、"
        "fcmp_transaction_historiesおよびフライト決済センター処理の正常完了を確認する。",
        "fcmp_transaction_historiesおよびフライト決済センター処理が全件正常完了",
        cases,
        value_hint="「同時N人」列に各段階の判定（OK/NG）を記録",
    )


def build_pt3(wb):
    base_id = len(FRONT_SCREENS_DISPLAY) + len(FRONT_SCREENS_CONCURRENT) + len(ADMIN_SCREENS) + 1
    large_defs = [
        ("お支払方法選択画面", "フロント",
         "1. 9サプライヤー×10商品=90商品の注文データを作成\n"
         "2. お支払方法選択画面へ遷移\n3. 画面表示・処理完了を確認\n"
         "4. orders, fcmp_transaction_historiesを確認\n5. k6で同時人数を1→2→3…と増やして繰り返す"),
        ("見積購入でのお支払方法選択画面", "フロント",
         "1. 見積購入フローで90商品の注文データを作成\n"
         "2. 見積購入の支払方法選択画面へ遷移\n3. 処理完了とDBを確認\n"
         "4. k6で同時人数を1→2→3…と増やして繰り返す"),
        ("出荷処理", "管理",
         "1. 90商品構成の出荷対象注文を用意\n2. 出荷処理を実行\n"
         "3. orders, fcmp_transaction_historiesを確認\n4. k6で同時人数を1→2→3…と増やして繰り返す"),
        ("注文変更処理", "管理",
         "1. 90商品構成の変更対象注文を用意\n2. 注文変更処理を実行\n"
         "3. 関連テーブル・フライト処理を確認\n4. k6で同時人数を1→2→3…と増やして繰り返す"),
        ("注文キャンセル処理", "管理",
         "1. 90商品構成のキャンセル対象注文を用意\n2. 注文キャンセル処理を実行\n"
         "3. 関連テーブル・フライト処理を確認\n4. k6で同時人数を1→2→3…と増やして繰り返す"),
    ]
    cases = []
    for i, (screen, site, steps) in enumerate(large_defs):
        cases.append((
            f"PT-{base_id + i:03d}", screen, site,
            LARGE_DATA_PRECONDITION,
            steps,
            "90商品/注文の大量データ条件下でも正常完了", "エラーなく正常完了", f"k6(1→10人)",
            *empty_ramp_tail(),
        ))

    build_detail_sheet(
        wb, "3_大量データの正常終了", 3,
        f"k6による負荷テスト。1注文あたり9サプライヤー×10商品=90商品の注文データを使用し、"
        "同時ユーザー数を1人→2人→3人…→10人と段階的に増やして確認する。",
        "orders, fcmp_transaction_histories等およびフライト決済センター処理が正常完了",
        cases,
        value_hint="「同時N人」列に各段階の判定（OK/NG）を記録",
    )


def build_pt4(wb):
    base_id = (
        len(FRONT_SCREENS_DISPLAY) + len(FRONT_SCREENS_CONCURRENT)
        + len(ADMIN_SCREENS) + 5 + 1
    )
    cases = [
        (
            f"PT-{base_id:03d}", "再オーソリバッチ", "バッチ",
            f"100注文分の再オーソリ対象データをfcmp_transaction_histories等に投入。{BATCH_STUB_NOTE}",
            "1. フライト決済センタースタブを有効化\n2. 再オーソリバッチを起動\n"
            "3. 全100注文に対して処理が実行されたことを確認\n4. 多重実行が発生していないことを確認",
            "100注文すべてが1サイクル内で正常処理", "多重実行なし・全件処理", "スタブ+データ投入",
            *empty_ramp_tail(),
        ),
        (
            f"PT-{base_id + 1:03d}", "障害取消実行バッチ（5分毎）", "バッチ",
            f"100注文分の障害取消対象データを投入。バッチは5分毎実行。{BATCH_STUB_NOTE}",
            "1. フライト決済センタースタブを有効化\n2. バッチ実行を待機または手動起動\n"
            "3. fcmp_transaction_historiesの全対象データが処理されたことを確認\n"
            "4. 5分を超える場合の挙動を確認",
            "実行サイクル内で処理完了、または5分超過時の想定挙動", "多重実行なし", "スタブ+データ投入",
            *empty_ramp_tail(),
            "5分以内に終わらない場合の挙動を必ず確認",
        ),
        (
            f"PT-{base_id + 2:03d}", "定期購入バッチ（契約サイト）", "契約サイト",
            "契約サイトに100件の定期購入データを投入。",
            "1. 定期購入バッチを起動\n2. 100件すべてが正常処理されたことを確認",
            "100件の定期購入が正常完了", "全件正常完了", "データ投入",
            *empty_ramp_tail(),
            "契約サイトで実施",
        ),
    ]
    build_detail_sheet(
        wb, "4_バッチ処理の大量データ", 4,
        "100注文分のデータを投入してバッチを実施。"
        "テストサイトのためフライト決済センター本体には負荷をかけず、スタブで代替する。",
        "短時間に大量処理対象が発生してもサイクル内に完了し、多重実行等の不具合がないこと",
        cases,
        value_hint="スタブ応答・バッチログ・fcmp_transaction_historiesの処理結果を記録",
    )


def build_env_sheet(wb):
    ws = wb.create_sheet("実施環境・ツール")
    rows = [
        ["項目", "内容", "記入欄"],
        ["テスト環境", "STG / 検証環境 等", ""],
        ["テスト実施期間", "", ""],
        ["手動テスト（項目1・2）", "同時1人→10人を段階的に実施（最大10名）", ""],
        ["負荷テストツール", "k6（テスト項目3で使用）", ""],
        ["k6スクリプト格納場所", "", ""],
        ["データ投入方法", "SQL / API / k6セットアップスクリプト 等", ""],
        ["大量データ条件", "1注文: 9サプライヤー × 10商品 = 90商品", ""],
        ["バッチテスト", "100注文。フライト決済センターはスタブで代替", ""],
        ["スタブ設定", "スタブのURL・切替方法・担当者", ""],
        ["監視対象", "APサーバ CPU/メモリ、DB、スタブ応答時間", ""],
        ["ログ確認先", "アプリログ、バッチログ、スタブログ", ""],
        ["テストデータ準備担当", "", ""],
        ["障害時エスカレーション", "", ""],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    style_header_row(ws, 1, 3)
    style_data_area(ws, 2, len(rows), 3)
    set_col_widths(ws, [24, 44, 30])


def main():
    wb = Workbook()
    build_overview(wb)
    build_ramp_sheet(wb)
    build_summary(wb)
    build_pt1(wb)
    build_pt2(wb)
    build_pt3(wb)
    build_pt4(wb)
    build_env_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
