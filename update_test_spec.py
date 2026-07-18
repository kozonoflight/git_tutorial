#!/usr/bin/env python3
"""アップロード版テスト仕様書を修正：エビデンス列追加 + v2内容 + 既存結果保持"""

import importlib.util
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/ubuntu/.cursor/projects/workspace/uploads/Primo_dataLayer________765c.xlsx"
OUT = "/workspace/Primo_dataLayer_テスト仕様書.xlsx"

# create_test_spec.py から定義を読み込み
spec = importlib.util.spec_from_file_location("cts", "/workspace/create_test_spec.py")
cts = importlib.util.module_from_spec(spec)
spec.loader.exec_module(cts)

HEADER_FILL = cts.HEADER_FILL
HEADER_FONT = cts.HEADER_FONT
NORMAL = cts.NORMAL
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def backup_results(wb):
    matrix = {}
    ws = wb["テストマトリクス"]
    for r in range(2, ws.max_row + 1):
        tc_id = ws.cell(r, 1).value
        if not tc_id:
            continue
        matrix[tc_id] = {
            "result": ws.cell(r, 6).value,
            "date": ws.cell(r, 7).value,
            "owner": ws.cell(r, 8).value,
            "note": ws.cell(r, 9).value,
        }

    details = {}
    for name in wb.sheetnames:
        if not name.startswith("TC-"):
            continue
        ws = wb[name]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            tc_id = ws.cell(r, 1).value
            if not tc_id:
                continue
            details[tc_id] = {
                "sheet": name,
                "result": ws.cell(r, headers.get("実施結果", 7)).value if "実施結果" in headers else None,
                "note": ws.cell(r, headers.get("備考", 8)).value if "備考" in headers else None,
                "actual": ws.cell(r, headers.get("実際の出力結果", 0)).value if "実際の出力結果" in headers else None,
                "evidence": ws.cell(r, headers.get("エビデンス", 0)).value if "エビデンス" in headers else None,
                # 手修正された期待結果・手順を保持
                "expected": ws.cell(r, headers.get("期待結果（コマンド出力）", 6)).value if "期待結果（コマンド出力）" in headers else None,
                "steps": ws.cell(r, headers.get("操作手順", 4)).value if "操作手順" in headers else None,
                "command": ws.cell(r, headers.get("確認コマンド（Console）", 5)).value if "確認コマンド（Console）" in headers else None,
            }
    return matrix, details


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_matrix_sheet(wb, saved_matrix):
    if "テストマトリクス" in wb.sheetnames:
        del wb["テストマトリクス"]
    ws = wb.create_sheet("テストマトリクス", 0)
    headers = [
        "テストID", "カテゴリ", "種別", "テスト項目", "対象イベント/機能", "優先度",
        "実施結果", "実施日", "実施者", "エビデンス", "備考",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [10, 14, 8, 38, 26, 8, 10, 12, 10, 24, 20])

    for r, tc in enumerate(cts.TEST_CASES, 2):
        tc_id, cat, name, target, pri, kind = tc
        saved = saved_matrix.get(tc_id, {})
        vals = [tc_id, cat, kind, name, target, pri,
                saved.get("result", ""), saved.get("date", ""), saved.get("owner", ""),
                saved.get("evidence", ""), saved.get("note", "")]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORMAL
            cell.border = BORDER
            cell.alignment = WRAP

    lr = len(cts.TEST_CASES) + 3
    ws.cell(row=lr, column=1, value="【凡例】").font = Font(bold=True, size=10)
    ws.cell(row=lr + 1, column=1, value="種別: 正常=成功時に送信 / 異常=失敗時・非該当時に送信しない")
    ws.cell(row=lr + 2, column=1, value="エビデンス: Consoleスクショ・GTMプレビュー・ページソースのファイル名を記載")
    ws.cell(row=lr + 3, column=1, value="実施結果: OK / NG / 未実施 / 保留")


def create_guide_sheet(wb):
    if "テスト実施手順（共通）" in wb.sheetnames:
        del wb["テスト実施手順（共通）"]
    cts.create_guide_sheet(wb)
    ws = wb["テスト実施手順（共通）"]
    # エビデンス手順を追記
    start = ws.max_row + 2
    rows = [
        ["■ エビデンス取得手順"],
        ["1. Console確認: コマンド実行結果が見えるようスクショを取得"],
        ["2. ページソース確認: Ctrl+U の該当箇所をスクショ、またはファイル名を記載"],
        ["3. GTMプレビュー: イベント発火画面をスクショ"],
        ["4. ファイル名例: TC-020_login_console.png / TC-002_noscript_source.png"],
        ["5. 詳細シートの「実際の出力結果」にConsole出力をコピペしても可"],
    ]
    for i, row in enumerate(rows):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=start + i, column=c, value=val)
            if val and val.startswith("■"):
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = NORMAL
            cell.alignment = WRAP


def create_detail_sheet(wb, sheet_name, category, saved_details):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    cases = [tc for tc in cts.TEST_CASES if tc[1] == category]
    headers = [
        "テストID", "種別", "テスト項目", "前提条件", "操作手順",
        "確認コマンド（Console）", "期待結果（コマンド出力）",
        "実際の出力結果", "エビデンス（スクショ/ファイル名）",
        "実施結果", "備考",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [10, 8, 24, 20, 24, 34, 34, 28, 22, 10, 16])

    row = 2
    for tc_id, _, name, _, priority, kind in cases:
        d = cts.DETAILS[tc_id]
        saved = saved_details.get(tc_id, {})
        # 手修正があれば優先（TC-002/003等）
        steps = saved.get("steps") or d["手順"]
        command = saved.get("command") or d["コマンド"]
        expected = saved.get("expected") or d["期待"]
        values = [
            tc_id, kind, f"【{priority}】{name}", d["前提"], steps, command, expected,
            saved.get("actual", ""), saved.get("evidence", ""),
            saved.get("result", ""), saved.get("note", ""),
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL
            cell.border = BORDER
            cell.alignment = WRAP
        ws.row_dimensions[row].height = 130
        row += 1
    ws.freeze_panes = "A2"


def main():
    src_wb = load_workbook(SRC)
    saved_matrix, saved_details = backup_results(src_wb)

    # 新規ブックとして再構築（古いTCシートは削除）
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    create_matrix_sheet(wb, saved_matrix)
    create_guide_sheet(wb)
    for category, sheet_name in cts.CATEGORY_SHEETS.items():
        create_detail_sheet(wb, sheet_name, category, saved_details)

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Test cases: {len(cts.TEST_CASES)}")
    preserved = sum(1 for tc in cts.TEST_CASES if saved_matrix.get(tc[0], {}).get("result"))
    print(f"Preserved matrix results: {preserved}")


if __name__ == "__main__":
    main()
